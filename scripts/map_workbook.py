from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = Path("workbook/Wind Rain Temp (Active).xlsx")
SHEETS = ["Wind", "Rain", "Temp", "Sum"]

def cell_value(cell):
    if cell.value is None:
        return ""
    return str(cell.value)

def main():
    wb = load_workbook(WORKBOOK_PATH, data_only=False)

    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        print("=" * 80)
        print(f"Sheet: {sheet_name}")
        print(f"Size: {ws.max_row} rows x {ws.max_column} columns")
        print()

        for row in ws.iter_rows():
            values = [cell_value(cell) for cell in row]
            if any(v != "" for v in values):
                trimmed = values[:20]
                print(f"Row {row[0].row:>3}: {trimmed}")

        print()

if __name__ == "__main__":
    main()
