from pathlib import Path
import csv
from openpyxl import load_workbook

WORKBOOK_PATH = Path("workbook/Wind Rain Temp (Active).xlsx")
OUTPUT_PATH = Path("data/weather_history.csv")

SHEET_CONFIG = {
    "Wind": {"variable": "wind", "unit": "m/s"},
    "Rain": {"variable": "rain", "unit": "mm/d"},
    "Temp": {"variable": "temp", "unit": "deg_c"},
}

MONTH_COLUMNS = {
    1: "D",
    2: "E",
    3: "F",
    4: "G",
    5: "H",
    6: "I",
    7: "J",
    8: "K",
    9: "L",
    10: "M",
    11: "N",
    12: "O",
}

DATA_START_ROW = 9
DATA_END_ROW = 65
YEAR_COLUMN = "C"


def main():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    rows = []

    for sheet_name, config in SHEET_CONFIG.items():
        ws = wb[sheet_name]

        for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
            year = ws[f"{YEAR_COLUMN}{row_idx}"].value

            if year is None:
                continue

            year = int(year)

            for month, col in MONTH_COLUMNS.items():
                value = ws[f"{col}{row_idx}"].value

                if value is None or value == "":
                    continue

                rows.append({
                    "year": year,
                    "month": month,
                    "date": f"{year}-{month:02d}-01",
                    "variable": config["variable"],
                    "value": value,
                    "unit": config["unit"],
                    "source_workbook": str(WORKBOOK_PATH),
                    "source_sheet": sheet_name,
                })

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with OUTPUT_PATH.open("w", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "year",
                "month",
                "date",
                "variable",
                "value",
                "unit",
                "source_workbook",
                "source_sheet",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote {len(rows)} rows to {OUTPUT_PATH}")

    latest = max((row["year"], row["month"]) for row in rows)
    print(f"Latest data point: {latest[0]}-{latest[1]:02d}")


if __name__ == "__main__":
    main()
