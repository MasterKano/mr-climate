from pathlib import Path

from openpyxl import load_workbook

ACTIVE_WORKBOOK = Path("workbook/Wind Rain Temp (Active).xlsx")
GENERATED_WORKBOOK = Path("workbook/Wind Rain Temp (Generated).xlsx")

SHEETS = ["Wind", "Rain", "Temp"]
DATA_START_ROW = 9
DATA_END_ROW = 65
DATA_START_COL = 4   # D
DATA_END_COL = 15    # O


def normalize(value):
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return round(float(value), 10)

    return value


def main():
    if not ACTIVE_WORKBOOK.exists():
        raise FileNotFoundError(f"Missing workbook: {ACTIVE_WORKBOOK}")

    if not GENERATED_WORKBOOK.exists():
        raise FileNotFoundError(
            f"Missing generated workbook: {GENERATED_WORKBOOK}. "
            "Run scripts/update_workbook_from_csv.py first."
        )

    active_wb = load_workbook(ACTIVE_WORKBOOK, data_only=True)
    generated_wb = load_workbook(GENERATED_WORKBOOK, data_only=True)

    mismatches = []

    for sheet_name in SHEETS:
        active_ws = active_wb[sheet_name]
        generated_ws = generated_wb[sheet_name]

        for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
            active_year = active_ws.cell(row=row_idx, column=3).value
            generated_year = generated_ws.cell(row=row_idx, column=3).value

            if active_year != generated_year:
                mismatches.append(
                    f"{sheet_name}!C{row_idx}: active year={active_year}, generated year={generated_year}"
                )

            for col_idx in range(DATA_START_COL, DATA_END_COL + 1):
                active_value = normalize(active_ws.cell(row=row_idx, column=col_idx).value)
                generated_value = normalize(generated_ws.cell(row=row_idx, column=col_idx).value)

                if active_value != generated_value:
                    cell = active_ws.cell(row=row_idx, column=col_idx).coordinate
                    mismatches.append(
                        f"{sheet_name}!{cell}: active={active_value}, generated={generated_value}"
                    )

    if mismatches:
        print("Comparison FAILED")
        print(f"Mismatches: {len(mismatches)}")
        for mismatch in mismatches[:50]:
            print("-", mismatch)
        if len(mismatches) > 50:
            print(f"... plus {len(mismatches) - 50} more")
        raise SystemExit(1)

    print("Comparison OK")
    print("Generated workbook matches active workbook weather data.")


if __name__ == "__main__":
    main()
