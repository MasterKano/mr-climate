from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook

INPUT_WORKBOOK = Path("workbook/Wind Rain Temp (Active).xlsx")
OUTPUT_WORKBOOK = Path("workbook/Wind Rain Temp (Generated).xlsx")
CSV_PATH = Path("data/weather_history.csv")

SHEET_CONFIG = {
    "wind": {"sheet": "Wind"},
    "rain": {"sheet": "Rain"},
    "temp": {"sheet": "Temp"},
}

MONTH_TO_COLUMN = {
    1: "D",
    2: "E",
    3: "F",
    4: "G",
    5: "H",
    6: "I",
    7: "J",
    8: "K",
    9: "L",
    10: "M",
    11: "N",
    12: "O",
}

YEAR_COLUMN = "C"
DATA_START_ROW = 9
DATA_END_ROW = 65


def clear_existing_values(ws):
    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in MONTH_TO_COLUMN.values():
            ws[f"{col}{row_idx}"] = None


def build_year_row_map(ws):
    year_to_row = {}

    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        year = ws[f"{YEAR_COLUMN}{row_idx}"].value

        if year is None:
            continue

        year_to_row[int(year)] = row_idx

    return year_to_row


def main():
    if not INPUT_WORKBOOK.exists():
        raise FileNotFoundError(f"Missing workbook: {INPUT_WORKBOOK}")

    if not CSV_PATH.exists():
        raise FileNotFoundError(f"Missing CSV: {CSV_PATH}")

    shutil.copy2(INPUT_WORKBOOK, OUTPUT_WORKBOOK)

    df = pd.read_csv(CSV_PATH)
    wb = load_workbook(OUTPUT_WORKBOOK)

    written_count = 0

    for variable, config in SHEET_CONFIG.items():
        ws = wb[config["sheet"]]
        clear_existing_values(ws)

        year_to_row = build_year_row_map(ws)
        variable_df = df[df["variable"] == variable]

        for record in variable_df.to_dict("records"):
            year = int(record["year"])
            month = int(record["month"])
            value = float(record["value"])

            if year not in year_to_row:
                raise ValueError(f"Year {year} not found in sheet {ws.title}")

            row_idx = year_to_row[year]
            col = MONTH_TO_COLUMN[month]
            ws[f"{col}{row_idx}"] = value
            written_count += 1

    wb.save(OUTPUT_WORKBOOK)

    print(f"Wrote {written_count} values")
    print(f"Saved generated workbook: {OUTPUT_WORKBOOK}")


if __name__ == "__main__":
    main()
