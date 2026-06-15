from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = Path("workbook/Wind Rain Temp (Active).xlsx")

def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH, data_only=False)

    print(f"Workbook: {WORKBOOK_PATH}")
    print()

    for ws in wb.worksheets:
        print(f"Sheet: {ws.title}")
        print(f"  Max row: {ws.max_row}")
        print(f"  Max column: {ws.max_column}")
        print(f"  Merged cells: {len(ws.merged_cells.ranges)}")
        print("  First 8 rows x 8 columns:")

        for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), min_col=1, max_col=min(8, ws.max_column)):
            values = [cell.value for cell in row]
            print(f"    {values}")

        print()

if __name__ == "__main__":
    main()
